# src/infrastructure/schedule_exporter.py

import pandas as pd
from pathlib import Path
from typing import Dict

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..scheduling.time_model import TimeModel


_COLOR_PALETTE = [
    "4CAF50", "2196F3", "FF9800", "9C27B0", "F44336",
    "009688", "E91E63", "3F51B5", "FF5722", "673AB7",
]

_HEADER_FILL  = PatternFill(start_color="1967D2", end_color="1967D2", fill_type="solid")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HOUR_FILL    = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
_THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# Grid constants (must match schedule_viewer_widget)
_GRID_STEP  = 30
_GRID_START = 7 * 60
_GRID_END   = 22 * 60
_GRID_ROWS  = (_GRID_END - _GRID_START) // _GRID_STEP


class ScheduleExporter:

    def __init__(self, time_model: TimeModel):
        self.time_model = time_model

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def to_excel(self, assignments: dict, output_path: str,
                 groups=None, course_name_by_code: dict = None,
                 include_grid: bool = True) -> None:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        name_map = self._build_name_map(assignments, groups, course_name_by_code)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            if include_grid:
                self._write_grid_sheets(writer, assignments, name_map)
            self._write_detail_sheet(writer, assignments, name_map)
            self._write_by_classroom_sheet(writer, assignments, name_map)

    def to_csv(self, assignments: dict, output_path: str,
               groups=None, course_name_by_code: dict = None) -> None:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        name_map = self._build_name_map(assignments, groups, course_name_by_code)
        df = self._detail_dataframe(assignments, name_map)
        df.to_csv(output_path, index=False, encoding="utf-8-sig")

    # ------------------------------------------------------------------
    # Detail sheet
    # ------------------------------------------------------------------

    def _write_detail_sheet(self, writer, assignments: dict, name_map: dict):
        df = self._detail_dataframe(assignments, name_map)
        df.to_excel(writer, sheet_name="Asignaciones", index=False)
        ws = writer.sheets["Asignaciones"]
        for cell in ws[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
        widths = [14, 35, 14, 14, 14, 12, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _detail_dataframe(self, assignments: dict, name_map: dict) -> pd.DataFrame:
        rows = []
        for gid, (cls, day, start_min, end_min) in sorted(assignments.items()):
            code = gid.rsplit("-G", 1)[0]
            group_num = gid.split("-P", 1)[0].rsplit("-G", 1)[1]
            rows.append({
                "Código Curso":  code,
                "Nombre Curso":  name_map.get(gid, ""),
                "Grupo":         f"{code}-G{group_num}",
                "Aula":          cls,
                "Día":           self.time_model.to_day_name(day),
                "Hora Inicio":   TimeModel.minutes_to_hhmm(start_min),
                "Hora Fin":      TimeModel.minutes_to_hhmm(end_min),
            })
        return pd.DataFrame(rows)

    # ------------------------------------------------------------------
    # By-classroom sheet
    # ------------------------------------------------------------------

    def _write_by_classroom_sheet(self, writer, assignments: dict, name_map: dict):
        rows = []
        for gid, (cls, day, start_min, end_min) in assignments.items():
            code = gid.rsplit("-G", 1)[0]
            group_num = gid.split("-P", 1)[0].rsplit("-G", 1)[1]
            rows.append({
                "Aula":         cls,
                "Código Curso": code,
                "Nombre Curso": name_map.get(gid, ""),
                "Grupo":        f"{code}-G{group_num}",
                "Día":          self.time_model.to_day_name(day),
                "Hora Inicio":  TimeModel.minutes_to_hhmm(start_min),
                "Hora Fin":     TimeModel.minutes_to_hhmm(end_min),
            })
        rows.sort(key=lambda r: (r["Aula"], r["Día"], r["Hora Inicio"]))
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name="Por Aula", index=False)
        ws = writer.sheets["Por Aula"]
        for cell in ws[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
        for i, w in enumerate([14, 14, 35, 14, 14, 12, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # Grid sheets (one per classroom)
    # ------------------------------------------------------------------

    def _write_grid_sheets(self, writer, assignments: dict, name_map: dict):
        # Group by classroom
        by_cls: dict[str, list] = {}
        for gid, (cls, day, start_min, end_min) in assignments.items():
            by_cls.setdefault(cls, []).append((gid, day, start_min, end_min))

        # Color map by course code
        course_codes = sorted(set(gid.rsplit('-G', 1)[0] for gid in assignments))
        course_colors = {
            code: _COLOR_PALETTE[i % len(_COLOR_PALETTE)]
            for i, code in enumerate(course_codes)
        }

        used: set[str] = set()
        for cls in sorted(by_cls):
            sheet_name = self._safe_sheet_name(f"Aula {cls}", used)
            used.add(sheet_name)
            self._write_single_grid(writer, sheet_name, cls,
                                    by_cls[cls], name_map, course_colors)

    def _write_single_grid(self, writer, sheet_name: str, classroom: str,
                           entries: list, name_map: dict, course_colors: dict):
        days = self.time_model.days
        n_rows = _GRID_ROWS
        n_cols = len(days) + 1

        # Build empty DataFrame
        time_labels = [TimeModel.minutes_to_hhmm(_GRID_START + r * _GRID_STEP)
                       for r in range(n_rows)]
        data = {d: [""] * n_rows for d in days}
        data = {"Hora": time_labels, **data}
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        ws = writer.sheets[sheet_name]

        # Style header
        for cell in ws[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER

        # Style hour column + empty cells
        empty_fill  = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
        center      = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for r in range(n_rows):
            excel_row = r + 2
            ws.cell(excel_row, 1).fill      = _HOUR_FILL
            ws.cell(excel_row, 1).font      = Font(bold=True, size=9)
            ws.cell(excel_row, 1).alignment = center
            ws.cell(excel_row, 1).border    = _THIN_BORDER
            for c in range(2, n_cols + 1):
                cell = ws.cell(excel_row, c)
                cell.fill      = empty_fill
                cell.alignment = center
                cell.border    = _THIN_BORDER

        # Place course blocks with merge
        occupied: dict[tuple, bool] = {}
        for gid, day, start_min, end_min in sorted(entries, key=lambda e: e[2]):
            day_name = self.time_model.to_day_name(day)
            if day_name not in days:
                continue
            col = days.index(day_name) + 2  # +2: 1-based + hour col

            start_row = (start_min - _GRID_START) // _GRID_STEP
            duration  = end_min - start_min
            span      = max(1, (duration + _GRID_STEP - 1) // _GRID_STEP)
            if start_row < 0 or start_row >= n_rows:
                continue
            span = min(span, n_rows - start_row)

            # Shrink if overlap
            for r in range(start_row, start_row + span):
                if (r, col) in occupied:
                    span = r - start_row
                    break
            if span < 1:
                continue
            for r in range(start_row, start_row + span):
                occupied[(r, col)] = True

            code      = gid.rsplit("-G", 1)[0]
            group_num = gid.split("-P", 1)[0].rsplit("-G", 1)[1]
            name      = name_map.get(gid, "")
            time_lbl  = f"{TimeModel.minutes_to_hhmm(start_min)}–{TimeModel.minutes_to_hhmm(end_min)}"
            text      = f"{code}\n{name}\nG{group_num}\n{time_lbl}" if name else f"{code}\nG{group_num}\n{time_lbl}"

            hex_color  = course_colors.get(code, _COLOR_PALETTE[0])
            course_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

            excel_row = start_row + 2
            cell = ws.cell(excel_row, col)
            cell.value     = text
            cell.fill      = course_fill
            cell.font      = Font(bold=True, size=9)
            cell.alignment = center
            cell.border    = _THIN_BORDER

            if span > 1:
                ws.merge_cells(
                    start_row=excel_row, start_column=col,
                    end_row=excel_row + span - 1, end_column=col
                )

        # Column widths / row heights
        ws.column_dimensions["A"].width = 7
        for c in range(2, n_cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 22
        ws.row_dimensions[1].height = 20
        for r in range(2, n_rows + 2):
            ws.row_dimensions[r].height = 38

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _build_name_map(self, assignments: dict, groups, course_name_by_code) -> dict:
        name_map = {}
        if groups:
            for g in groups:
                if g.course_name:
                    name_map[g.group_id] = g.course_name
        course_name_by_code = course_name_by_code or {}
        for gid in assignments:
            if gid not in name_map:
                code = gid.rsplit("-G", 1)[0]
                if code in course_name_by_code:
                    name_map[gid] = course_name_by_code[code]
        return name_map

    def _safe_sheet_name(self, name: str, used: set) -> str:
        for ch in r'\/*?:[]':
            name = name.replace(ch, "-")
        name = name[:31]
        if name not in used:
            return name
        i = 2
        while True:
            candidate = f"{name[:28]}_{i}"
            if candidate not in used:
                return candidate
            i += 1
